#CreditPro
import os

from unipath import Path
from pandas import ExcelWriter
from openpyxl import load_workbook
import pandas.io.formats.excel
import pandas as pd
import openpyxl
def matrizBases(file_list):   #Función de la matriz, list_file son las direcciones de las bases de datos.
    matrices = []
    dic_princ = []
    for file_name in file_list:
        matriz = pd.DataFrame({'TIPO DE IDENTIFICACION':[], 'CEDULA - RUC':[], 'NOMBRE APELLIDO':[], 'DIRECCION':[], 'SECTOR':[],'CANTON':[], 'PROVINCIA':[], 'TELEFONO':[], 'FORMA DE PAGO':[],
                               'CASADO':[], 'TRABAJA RELACION DEPENDENCIA':[], 'TRABAJA SIN RELACION DEPENDENCIA':[], 'INGRESOS':[], 'EDAD':[], 'SEXO':[], 'VEHICULO PROPIO':[], 'CASA PROPIA':[]})                  #Diccionario vacío que se usará  para crear la matriz con los parámetros
        print(file_name)
        dic_comparacion = pd.read_excel(file_name).loc[0].to_dict()             #Selecciona las etiquetas y la primera fila de datos con loc[0] y los convierte en un diccionario con clave "etiquetas" y valor el elemento de cada fila.
        for etiqueta in dic_comparacion:                                        #Iteramos sobre cada clave(etiqueta) del diccionario.
            columna_etiqueta = pd.read_excel(file_name)[etiqueta]               #De la base de datos principal, seleccionamos cada una de sus columnas.
            if str(etiqueta).upper().startswith('CED') or str(etiqueta).upper().startswith('RUC') or str(etiqueta).upper().startswith('PAS') : #Comparamos con upper si la etiqueta empieza con "CED" de cédula o "RUC" de ruc o "PAS" de pasaporte. El uper es importante porque puede que el dato esté en minúsculas.
                columna_etiqueta2 = []
                for filt_ced in columna_etiqueta:
                    if len(str(filt_ced)) == 9:
                        columna_etiqueta2.append('0'+ str(filt_ced))
                    else:
                        columna_etiqueta2.append(str(filt_ced))
                matriz['CEDULA - RUC'] = columna_etiqueta2                               #Si cumple la condición anterior sobre la etiqueta, agregamos toda esa columna de identificaciones del archivo en la columna "CÉDULA - RUC" de la matriz principal.
                lista_cedoruc = []                                                         #Creamos una lista vacía donde agregaremos el tipo de identificación.
                for cedoruc in columna_etiqueta:                                           #Recorre cada número de identificación.
                        if len(str(cedoruc)) == 10 or len(str(cedoruc)) == 9:                                             #Pregunta si el número de identificación tiene 10 caracteres.
                            lista_cedoruc.append('C')                                           #Si los tiene, se agrega "C" de cédula en la lista "lista_cedoruc".
                        elif len(str(cedoruc)) == 13:                                           #Pregunta si el número de identificación tiene 13 caracteres.
                            lista_cedoruc.append('R')                                           #Si los tiene, se agrega "R" de RUC a la lista "lista_cedoruc".
                        elif str(cedoruc).endswith('001'):                                      #Pregunta si termina en 001 para hacer otra validación sobre el RUC.
                            lista_cedoruc.append('R')                                           #Si no tiene 13 caracteres pero termina en 001, se considera RUC y se añade "R" de RUC a la lista.
                        elif len(str(cedoruc)) == 7:                                            #Si tiene 7 caracteres.
                            lista_cedoruc.append('P')                                           #Por tanto, aquí se agrega la "P" de pasaporte en la lista "lista_cedoruc"
                        else:                                                                   #Si tiene cualquier otra naturaleza, se dice que es un pasaporte.
                            lista_cedoruc.append('nan')                                         #Se agrega "NaN" que se refiere a que no lo reconoce o es incorrecto para la base.

                matriz['TIPO DE IDENTIFICACION'] = lista_cedoruc                            #Una vez recorrido todo el archivo en esa columna, la lista "lista_cedoruc" con los datos sobre si es RUC, Cédula o pasaporte, se añade a la matriz principal en la columna "CEDULA - RUC".
            elif str(etiqueta).upper().startswith('NOM') or str(etiqueta).upper().startswith('APEL'): #Se pregunta si la etiqueta empieza con "NOM" de nombre o si empieza con "APEL" de apellido.
                matriz['NOMBRE APELLIDO'] = columna_etiqueta                                #Si cumple la condición anterior, simplemente se añade toda esa columna de nombres y apellidos a la matriz principal en la columna "NOMBRE Y APELLIDO".
            elif str(etiqueta).upper().startswith('DIRE'):                                       #Se pregunta si la etiqueta empieza con "DIRE" de dirección, no se le agrego más de la palabra ya que podría generar un conflicto si tiene tilde o no en el documento que se está filtrando.
                matriz['DIRECCION'] = columna_etiqueta                                      #Si cumple, se añade la columna de la etiqueta en la matriz principal en la columna "DIRECCION".
            elif str(etiqueta).upper().startswith('SECT') or str(etiqueta).upper().startswith('ZONA'):#Se pregunta si la etiqueta empieza con "SECT" o con "ZONA".
                matriz['SECTOR'] = columna_etiqueta                                            #Si cumple, se añade la columna de la etiqueta en la matriz principal en la columna "SECTOR".
            elif str(etiqueta).upper().startswith('CANT') or str(etiqueta).upper().startswith('CIUDAD'): #Se pregunta si la etiqueta empieza con "CANT" de cantón o "CIUDAD", a cantón no se le escribió completo proque podría generar un conflicto con la tilde.
                matriz['CANTON'] = columna_etiqueta                                             #Si cumple, se añade la columna de la etiqueta a la matriz principal en la columna "CANTON"
            elif str(etiqueta).upper().startswith('PROVINCIA'):                                      #Se pregunta si la etiqueta es "PROVINCIA"
                matriz['PROVINCIA'] = columna_etiqueta                                             #Si lo es, que se añada la columna de la etiqueta en la matriz principal en la columna "PROVINCIA"
            elif str(etiqueta).upper().startswith('CEL') or str(etiqueta).upper().startswith('TEL') or str(etiqueta).upper().startswith('CON'): #Se filtra con "CEL" de celular, "TEL" de teléfono o "CON" de convencional para obtener el número de teléfono.
                 matriz['TELEFONO'] = columna_etiqueta                                                                          #Si la etiqueta cumple la condición, que agregue esa columna de la etiqueta en la columna "TELEFONO" de la matriz principal.
            elif str(etiqueta).upper().startswith('FORMA DE PAGO') or str(etiqueta).upper().startswith('FOR') or str(etiqueta).upper().startswith('TIPO') or str(etiqueta).upper().startswith('MODO') or str(etiqueta).upper().startswith('MODO DE PAGO'):  #Se filtra la etiqueta para forma de pago y se usan palabras con las cuales pueda ser identificada esa etiqueta.
                matriz['FORMA DE PAGO'] = columna_etiqueta                                                  #Se agrega la columna de etiqueta a la matriz princiapl en la columna "FORMA DE PAGO" si se cumplen las validaciones anteriores.
            elif str(etiqueta).upper().startswith('CASAD') or str(etiqueta).upper().startswith('ESTADO CIVIL') or str(etiqueta).upper().startswith('CIV'):     #Se pregunta sobre su estado civil de distintas maneras, como no se sabe si es hombre o mujer, se pregunta "CASAD" ya que puede estar casado o casada la persona.
                list_est_civ = []                                   #Creamos una lista vacía donde se van a añadir si están o no casados o casadas cada cliente.
                for est_civ in columna_etiqueta:                    #Iteramos sobre cada elemento de la columna que es del estado civil.
                    if str(etiqueta).upper().startswith('CASAD'):         #Puede estar de distintas maneras el estado civil, en este caso se pregunta si la etiqueta empieza con "CASAD" por si de esa manera está en el documento el dato.
                        if str(est_civ).upper().startswith('Y') or str(est_civ).upper().startswith('S'): #Se pregunta si dice "Y" de yes o "S" de sí.
                            list_est_civ.append('S')                            #Si se cumple, se añade "S" a la lista "list_est_civ".
                        elif str(est_civ).upper() != 'NAN':                     #Esto para no confundir con un dato errante.
                            if str(est_civ).upper().startswith('N'):                                                 #Si dice que "N" de no anda casado.
                                list_est_civ.append('N')                            #Que ponga "N" en la lista "list_est_civ".
                        elif str(est_civ).upper() == 'NAN':
                            list_est_civ.append('nan')
                        else:                                                   #Si no dice nada de lo anterior o el campo está vacío...
                            list_est_civ.append('nan')                           #Que añada "Na" de no se sabe.
                    elif str(etiqueta).upper().startswith('CIV') or str(etiqueta).upper().startswith('ESTADO CIVIL'):   #La etiqueta del documento puede decir "ESTADO CIVIL" por tanto lo filtramos.
                        if str(est_civ).upper().startswith('CASAD'):                           #Si en estado civil está "CASAD" de casado o casa, se cumple.
                            list_est_civ.append('S')                                #Si cumple, se agrega "S" en la lista "list_est_civ".
                        elif str(est_civ).upper().startswith('SOLTER') or str(est_civ).upper().startswith('DIVORCIAD') or str(est_civ).upper().startswith('VIUD'):                           #Si no está casado, es decir, su estado civil es cualquier otro-
                            list_est_civ.append('N')                                #Se añade "N" a la lista "list_est_civ".
                        else:                                                       #Si no es algún estado civil.
                            list_est_civ.append('nan')                               #Se pone "Na" de que no se sabe o no hay información.
                matriz['CASADO'] = list_est_civ                                     #Una vez que se haya llenado la lista, agregamos a la matriz principal en la columna "CASADO".
            elif str(etiqueta).upper().startswith('TRABAJA RELACION DEPENDENCIA'):                     #Preguntamos si la etiqueta es "NEGOCIO PROPIO"
                list_neg_prop = []                                                  #Creamos una lista que se llena con respecto a si tiene negocio propio
                list_no_neg_prop = []                                               #Lo contrario a la lista anterior
                for neg_prop in columna_etiqueta:                                   #Se itera cada elemento de la etiqueta que es practicamente si tiene negocio propio
                    if str(neg_prop).upper().startswith('Y') or str(neg_prop).upper().startswith('S'): #Se pregunta Si la respuesta sobre el negocio propio es Sí.
                        list_neg_prop.append('S')                                              #Si cumple, se agrega "S" de sí, en la lista de negocios propios y "N" de no, en el negocio no propio.
                        list_no_neg_prop.append('N')
                    elif str(neg_prop).upper() == 'N' or str(neg_prop).upper() == 'NO':
                        list_neg_prop.append('N')                                               #Si no tiene, se agrega en la lista "list_neg_prop" "N" de que no tiene.
                        list_no_neg_prop.append('S')                                            #Aquí simplemente es lo contrario.
                    elif str(neg_prop).upper() != 'N' or str(neg_prop).upper() != 'NO' or str(neg_prop).upper() != 'S' or str(neg_prop).upper() != 'SI': #Si dice algo diferente a Sí y No, entonces ponemos "Na" en ambas listas de que no hay información o que no lo reconoce.
                        list_neg_prop.append('nan')
                        list_no_neg_prop.append('nan')
                matriz['TRABAJA RELACION DEPENDENCIA'] = list_neg_prop       #Se agrega la lista de si no tiene negocios propios en la columna de dependencia de la matriz principal.
                matriz['TRABAJA SIN RELACION DEPENDENCIA'] = list_no_neg_prop      #Se agrega la lista de si tiene negocios propios  en la columna de independencia de la matriz principal.
            elif str(etiqueta).upper().startswith('INGRESOS') or str(etiqueta).upper().startswith('SALARIO') or str(etiqueta).upper().startswith('GANANCIA'):  #Para saber los salarios, se usa el filtro de etiqueta de "SALARIO" o "GANANCIA"
                matriz['INGRESOS'] = columna_etiqueta                       #Si cumple lo anterior, que se agrega toda la columna de la etiqueta ingresos en la columna de la matriz principal "INGRESOS"
            elif str(etiqueta).upper().startswith('EDAD') or str(etiqueta).upper().startswith('AGE'):   #Se pregunta si la etiqueta se refiere a la edad del usuario, si lo es.
                matriz['EDAD'] = columna_etiqueta                                           #Agregamos la columna de la etiqueta en la columna "EDAD" de la matriz principal.
            elif str(etiqueta).upper().startswith('GENERO') or str(etiqueta).upper().startswith('GÉNERO') or str(etiqueta).upper().startswith('SEXO') or str(etiqueta).upper().startswith('GENDER'): #Se filtra el género o sexo del archivo.
                list_sexo = []                      #Se crea una lista sobre el tipo de sexo que son.
                for sexo in columna_etiqueta:       #Se itera en la columna de la etiqueta para obtener el género de cada cliente.
                    if str(sexo).upper().startswith('MASCULINO') or str(sexo).upper().startswith('M') : #Si es masculino.
                        list_sexo.append('M')                   #Si cumple, añade "M" a la lista "list_sexo".
                    elif str(sexo).upper().startswith('FEMENINO') or str(sexo).upper().startswith('F'): #Si es femenino.
                        list_sexo.append('F')   #Se añade "F" a la lista "list_sexo"
                    else:                       #Si no está masculino o femenino, es porque hay un dato erróneo o está vacío el espacio.
                        list_sexo.append('nan')      #Por tanto, se añade Na.
                matriz['SEXO'] = list_sexo          #Se agrega la lista llena de "list_sexo" a la matriz principal en la columna "SEXO"
            elif str(etiqueta).upper().startswith('VEHICULO PROPIO') or str(etiqueta).upper().startswith('CARRO') or str(etiqueta).upper().startswith('VEH'): #Se filtra la etiqueta para saber si se refiere al vehículo propio del cliente.
                list_vehi_prop = []     #Se crea una lista vacía para saber si el cliente tiene o no vehículos propios.
                for vehi_prop in columna_etiqueta:  #Se itera cada elemento de la columna de vehículos.
                    if str(vehi_prop).upper().startswith('Y') or str(vehi_prop).upper().startswith('S'): #Si la información cumple esto.
                        list_vehi_prop.append('S')          #Se añade "S" de sí a la lista "list_vehi_prop".
                    elif str(vehi_prop).upper() != 'NAN':
                        if str(vehi_prop).upper().startswith('N'): #Si la información cumple esto.
                            list_vehi_prop.append('N')           #Se añade "N" de no a la lista "list_vehi_prop".
                    else:                                   #Si está vacío o tiene información ambigua al programa.
                        list_vehi_prop.append('nan')          #Que añada "Na" a la lista "list_vehi_prop"
                matriz['VEHICULO PROPIO'] = list_vehi_prop      #Se agrega la lista completa "list_vehi_prop" a la columna "VEHICULO PROPIO" de la matriz principal.
            elif str(etiqueta).upper().startswith('CASA PROPIA') or str(etiqueta).upper().startswith('VIVIENDA PROPIA'):  #Si la etiqueta tiene que ver con casa propia
                list_viv_prop = []              #Se crea una lista vacía donde se añadirá si tiene o no casa propia.
                for viv_prop in columna_etiqueta:   #Itera sobre cada elemento de la columna de la etiqueta, que es sí o no.
                    if str(viv_prop).upper().startswith('Y') or str(viv_prop).upper().startswith('S'):        #Si cumple la condición de que tiene casa propia.
                        list_viv_prop.append('S')               #Que añada "S" a la lista "list_viv_prop"
                    elif str(viv_prop).upper() != 'NAN':
                        if str(viv_prop).upper().startswith('N'):      #Si cumple que tiene "N" de no.
                            list_viv_prop.append('N')               #Añade "N" a la lista "list_viv_prop"
                    else:                                       #Si la información es ambigua al programa o está vacío.
                        list_viv_prop.append('nan')              #Se añade "Na" a la lista.
                matriz['CASA PROPIA'] = list_viv_prop           #Se añade la lista completa con la información de la casa propia.
        if len(dic_princ) == 0:
            dic_princ.append(matriz.to_dict())
        else:
            matrices.append(matriz.to_dict())                        #Convertirmos cada matriz en diccionario y los agregamos a una lista "matrices"
    if len(dic_princ) == 1 and len(matrices) == 0:
        return pd.DataFrame(dic_princ[0])
    elif len(matrices) > 0:
        dic1 = dic_princ[0]
        dic_1_tip = list(dic1['TIPO DE IDENTIFICACION'].values())
        dic_1_ced = list(dic1['CEDULA - RUC'].values())
        dic_1_nya = list(dic1['NOMBRE APELLIDO'].values())
        dic_1_dir = list(dic1['DIRECCION'].values())
        dic_1_sec = list(dic1['SECTOR'].values())
        dic_1_can = list(dic1['CANTON'].values())
        dic_1_pro = list(dic1['PROVINCIA'].values())
        dic_1_tel = list(dic1['TELEFONO'].values())
        dic_1_for = list(dic1['FORMA DE PAGO'].values())
        dic_1_cas = list(dic1['CASADO'].values())
        dic_1_dep = list(dic1['TRABAJA RELACION DEPENDENCIA'].values())
        dic_1_ind = list(dic1['TRABAJA SIN RELACION DEPENDENCIA'].values())
        dic_1_ing = list(dic1['INGRESOS'].values())
        dic_1_age = list(dic1['EDAD'].values())
        dic_1_sex = list(dic1['SEXO'].values())
        dic_1_veh = list(dic1['VEHICULO PROPIO'].values())
        dic_1_casprop = list(dic1['CASA PROPIA'].values())
        nuevo_dic_1 = {'TIPO DE IDENTIFICACION': dic_1_tip, 'CEDULA - RUC': dic_1_ced, 'NOMBRE APELLIDO': dic_1_nya, 'DIRECCION': dic_1_dir, 'SECTOR': dic_1_sec, 'CANTON': dic_1_can, 'PROVINCIA': dic_1_pro, 'TELEFONO':dic_1_tel, 'FORMA DE PAGO': dic_1_for, 'CASADO': dic_1_cas, 'TRABAJA RELACION DEPENDENCIA': dic_1_dep, 'TRABAJA SIN RELACION DEPENDENCIA': dic_1_ind, 'INGRESOS': dic_1_ing, 'EDAD': dic_1_age, 'SEXO': dic_1_sex, 'VEHICULO PROPIO': dic_1_veh, 'CASA PROPIA': dic_1_casprop}
        for dic in matrices:
            dic_n_ced = list(dic['CEDULA - RUC'].values())
            dic_n_tip = list(dic['TIPO DE IDENTIFICACION'].values())
            dic_n_nya = list(dic['NOMBRE APELLIDO'].values())
            dic_n_dir = list(dic['DIRECCION'].values())
            dic_n_sec = list(dic['SECTOR'].values())
            dic_n_can = list(dic['CANTON'].values())
            dic_n_pro = list(dic['PROVINCIA'].values())
            dic_n_tel = list(dic['TELEFONO'].values())
            dic_n_for = list(dic['FORMA DE PAGO'].values())
            dic_n_cas = list(dic['CASADO'].values())
            dic_n_dep = list(dic['TRABAJA RELACION DEPENDENCIA'].values())
            dic_n_ind = list(dic['TRABAJA SIN RELACION DEPENDENCIA'].values())
            dic_n_ing = list(dic['INGRESOS'].values())
            dic_n_age = list(dic['EDAD'].values())
            dic_n_sex = list(dic['SEXO'].values())
            dic_n_veh = list(dic['VEHICULO PROPIO'].values())
            dic_n_casprop = list(dic['CASA PROPIA'].values())
            nuevo_dic_n = {'TIPO DE IDENTIFICACION': dic_n_tip, 'CEDULA - RUC': dic_n_ced, 'NOMBRE APELLIDO': dic_n_nya, 'DIRECCION': dic_n_dir, 'SECTOR': dic_n_sec, 'CANTON': dic_n_can, 'PROVINCIA': dic_n_pro, 'TELEFONO':dic_n_tel, 'FORMA DE PAGO': dic_n_for, 'CASADO': dic_n_cas, 'TRABAJA RELACION DEPENDENCIA': dic_n_dep,'TRABAJA SIN RELACION DEPENDENCIA': dic_n_ind, 'INGRESOS': dic_n_ing, 'EDAD': dic_n_age, 'SEXO': dic_n_sex, 'VEHICULO PROPIO': dic_n_veh, 'CASA PROPIA': dic_n_casprop}
            for ced in nuevo_dic_n['CEDULA - RUC']:
                if str(ced) not in nuevo_dic_1['CEDULA - RUC']:
                    nuevo_dic_1['CEDULA - RUC'].append(str(ced))
                    ind = nuevo_dic_n['CEDULA - RUC'].index(str(ced))
                    nuevo_dic_1['TIPO DE IDENTIFICACION'].append(str(nuevo_dic_n['TIPO DE IDENTIFICACION'][ind]))
                    nuevo_dic_1['NOMBRE APELLIDO'].append(str(nuevo_dic_n['NOMBRE APELLIDO'][ind]))
                    nuevo_dic_1['DIRECCION'].append(str(nuevo_dic_n['DIRECCION'][ind]))
                    nuevo_dic_1['SECTOR'].append(str(nuevo_dic_n['SECTOR'][ind]))
                    nuevo_dic_1['CANTON'].append(str(nuevo_dic_n['CANTON'][ind]))
                    nuevo_dic_1['PROVINCIA'].append(str(nuevo_dic_n['PROVINCIA'][ind]))
                    nuevo_dic_1['TELEFONO'].append(str(nuevo_dic_n['TELEFONO'][ind]))
                    nuevo_dic_1['CASADO'].append(str(nuevo_dic_n['CASADO'][ind]))
                    nuevo_dic_1['FORMA DE PAGO'].append(str(nuevo_dic_n['FORMA DE PAGO'][ind]))
                    nuevo_dic_1['TRABAJA RELACION DEPENDENCIA'].append(str(nuevo_dic_n['TRABAJA RELACION DEPENDENCIA'][ind]))
                    nuevo_dic_1['TRABAJA SIN RELACION DEPENDENCIA'].append(str(nuevo_dic_n['TRABAJA SIN RELACION DEPENDENCIA'][ind]))
                    nuevo_dic_1['INGRESOS'].append(str(nuevo_dic_n['INGRESOS'][ind]))
                    nuevo_dic_1['EDAD'].append(str(nuevo_dic_n['EDAD'][ind]))
                    nuevo_dic_1['SEXO'].append(str(nuevo_dic_n['SEXO'][ind]))
                    nuevo_dic_1['VEHICULO PROPIO'].append(str(nuevo_dic_n['VEHICULO PROPIO'][ind]))
                    nuevo_dic_1['CASA PROPIA'].append(str(nuevo_dic_n['CASA PROPIA'][ind]))
                elif str(ced) in nuevo_dic_1['CEDULA - RUC']:
                    indn = nuevo_dic_n['CEDULA - RUC'].index(ced)
                    ind1 = nuevo_dic_1['CEDULA - RUC'].index(ced)
                    if str(nuevo_dic_1['NOMBRE APELLIDO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['NOMBRE APELLIDO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['NOMBRE APELLIDO'][ind1] = nuevo_dic_n['NOMBRE APELLIDO'][indn]
                    if str(nuevo_dic_1['DIRECCION'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['DIRECCION'][indn]).upper() != 'NAN':
                        nuevo_dic_1['DIRECCION'][ind1] = nuevo_dic_n['DIRECCION'][indn]
                    if str(nuevo_dic_1['SECTOR'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['SECTOR'][indn]).upper() != 'NAN':
                        nuevo_dic_1['SECTOR'][ind1] = nuevo_dic_n['SECTOR'][indn]
                    if str(nuevo_dic_1['CANTON'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['CANTON'][indn]).upper() != 'NAN':
                        nuevo_dic_1['CANTON'][ind1] = nuevo_dic_n['CANTON'][indn]
                    if str(nuevo_dic_1['PROVINCIA'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['PROVINCIA'][indn]).upper() != 'NAN':
                        nuevo_dic_1['PROVINCIA'][ind1] = nuevo_dic_n['PROVINCIA'][indn]
                    if str(nuevo_dic_1['TELEFONO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['TELEFONO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['TELEFONO'][ind1] = nuevo_dic_n['TELEFONO'][indn]
                    if str(nuevo_dic_1['FORMA DE PAGO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['FORMA DE PAGO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['FORMA DE PAGO'][ind1] = nuevo_dic_n['FORMA DE PAGO'][indn]
                    if str(nuevo_dic_1['CASADO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['CASADO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['CASADO'][ind1] = nuevo_dic_n['CASADO'][indn]
                    if str(nuevo_dic_1['TRABAJA RELACION DEPENDENCIA'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['TRABAJA RELACION DEPENDENCIA'][indn]).upper() != 'NAN':
                        nuevo_dic_1['TRABAJA RELACION DEPENDENCIA'][ind1] = nuevo_dic_n['TRABAJA RELACION DEPENDENCIA'][indn]
                    if str(nuevo_dic_1['TRABAJA SIN RELACION DEPENDENCIA'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['TRABAJA SIN RELACION DEPENDENCIA'][indn]).upper() != 'NAN':
                        nuevo_dic_1['TRABAJA SIN RELACION DEPENDENCIA'][ind1] = nuevo_dic_n['TRABAJA SIN RELACION DEPENDENCIA'][indn]
                    if str(nuevo_dic_1['INGRESOS'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['INGRESOS'][indn]).upper() != 'NAN':
                        nuevo_dic_1['INGRESOS'][ind1] = nuevo_dic_n['INGRESOS'][indn]
                    if str(nuevo_dic_1['EDAD'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['EDAD'][indn]).upper() != 'NAN':
                        nuevo_dic_1['EDAD'][ind1] = nuevo_dic_n['EDAD'][indn]
                    if str(nuevo_dic_1['SEXO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['SEXO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['SEXO'][ind1] = nuevo_dic_n['SEXO'][indn]
                    if str(nuevo_dic_1['VEHICULO PROPIO'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['VEHICULO PROPIO'][indn]).upper() != 'NAN':
                        nuevo_dic_1['VEHICULO PROPIO'][ind1] = nuevo_dic_n['VEHICULO PROPIO'][indn]
                    if str(nuevo_dic_1['CASA PROPIA'][ind1]).upper() == 'NAN' and str(nuevo_dic_n['CASA PROPIA'][indn]).upper() != 'NAN':
                        nuevo_dic_1['CASA PROPIA'][ind1] = nuevo_dic_n['CASA PROPIA'][indn]
        matriz_filtrada = pd.DataFrame(nuevo_dic_1)

        return matriz_filtrada
#dataFrameJT= matrizBases(["prueba matriz.xlsx","prueba matriz2.xlsx","prueba matriz3.xlsx"])
#print(dataFrameJT)
def parametrosDefault(curpath):
    #print(curpath)
    f=open(curpath+"\\"+"DATOS PREDETERMINADOS.txt","r")
    for i in f:
        separar=i.split(",")
        return separar

def funcionPuntajes(dataFrameJT,lista_param,stringpath):
    #funcion que va a analizar el dataframe de Joel Torres
    #print(os.getcwd())
    print(stringpath)
    directorio=os.path.dirname(stringpath)
    os.chdir(directorio)
    print(os.getcwd())
    # defaultlistaparam = [5,10,0,10,10,5,10,5,10,5,5,7,9,10,5,10,10,5,10,5,4,5,7,9,10,10,5]
    numeroFilas = dataFrameJT.shape[1]  # variable que me va a dar el numero de filas
    puntaje = []
    lista_param_numerica = lista_param
    listaParametros = parametrosDefault(directorio)

    for i in range(0, numeroFilas + 1):  # recorro las filas n veces, comenzando por el indice 0 hasta el indice n
        listaConteo = []  # un contador que comienza en 0
        nFila = dataFrameJT.ix[i:i]  # comienza la lectura de la primera fila, luego segunda, luego tercera, etc etc...
        listaDeListaDeFilas = nFila.values
        for k in listaDeListaDeFilas:
            listaDeFila = k
            identificacion = listaDeFila[0]  # sacando parametros
            formaPago = listaDeFila[8]
            casado = listaDeFila[9]
            trabajaConDependencia = listaDeFila[10]
            negocioPropio = listaDeFila[11]
            ingresos = float(listaDeFila[12])
            edad = float(listaDeFila[13])
            sexo = listaDeFila[14]
            vehiculoPropio = listaDeFila[15]
            casaPropia = listaDeFila[16]
            if identificacion == "C":  # Comienzo a poner mis condiciones
                if lista_param_numerica[0] == None or lista_param_numerica[0].isnumeric()==False :
                    #print("Se asigna por defecto")
                    listaConteo.append(listaParametros[0])
                else:
                    listaConteo.append(lista_param_numerica[0])
            if identificacion == "R":
                if lista_param_numerica[1] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[1])
                else:
                    listaConteo.append(lista_param_numerica[1])
            if identificacion == "P":
                if lista_param_numerica[2] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[2])
                else:
                    listaConteo.append(lista_param_numerica[2])
            if formaPago == "TARJETA":
                if lista_param_numerica[3] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[3])
                else:
                    listaConteo.append(lista_param_numerica[3])
            if formaPago == "BANCO":
                if lista_param_numerica[4] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[4])
                else:
                    listaConteo.append(lista_param_numerica[4])
                # listaConteo.append(10)
            if formaPago == "COOPERATIVA":
                if lista_param_numerica[5] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[5])
                else:
                    listaConteo.append(lista_param_numerica[5])
                # listaConteo.append(5)
            if casado == "S":
                if lista_param_numerica[6] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[6])
                else:
                    listaConteo.append(lista_param_numerica[6])
                # listaConteo.append(10)
            if casado == "N":
                if lista_param_numerica[7] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[7])
                else:
                    listaConteo.append(lista_param_numerica[7])
                # listaConteo.append(5)
            if negocioPropio == "S":
                if lista_param_numerica[8] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[8])
                else:
                    listaConteo.append(lista_param_numerica[8])
                # listaConteo.append(10)
            if negocioPropio == "N":
                if lista_param_numerica[9] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[9])
                else:
                    listaConteo.append(lista_param_numerica[9])
                # listaConteo.append(5)
            if ingresos >= 374 and ingresos <= 500:
                if lista_param_numerica[10] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[10])
                else:
                    listaConteo.append(lista_param_numerica[10])
                # listaConteo.append(5)
            if ingresos >= 501 and ingresos <= 700:
                if lista_param_numerica[11] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[11])
                else:
                    listaConteo.append(lista_param_numerica[11])
                # listaConteo.append(7)
            if ingresos >= 701 and ingresos <= 900:
                if lista_param_numerica[12] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[12])
                else:
                    listaConteo.append(lista_param_numerica[12])
                # listaConteo.append(9)
            if ingresos >= 901:
                if lista_param_numerica[13] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[13])
                else:
                    listaConteo.append(lista_param_numerica[13])
                # listaConteo.append(10)
            if sexo == "M":
                if lista_param_numerica[14] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[14])
                else:
                    listaConteo.append(lista_param_numerica[14])
                # listaConteo.append(5)
            if sexo == "F":
                if lista_param_numerica[15] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[15])
                else:
                    listaConteo.append(lista_param_numerica[15])
                # listaConteo.append(10)
            if vehiculoPropio == "S":
                if lista_param_numerica[16] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[16])
                else:
                    listaConteo.append(lista_param_numerica[16])
                # listaConteo.append(10)
            if vehiculoPropio == "N":
                if lista_param_numerica[17] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[17])
                else:
                    listaConteo.append(lista_param_numerica[17])
                # listaConteo.append(5)
            if casaPropia == "S":
                if lista_param_numerica[18] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[18])
                else:
                    listaConteo.append(lista_param_numerica[7])
                # listaConteo.append(10)
            if casaPropia == "N":
                if lista_param_numerica[19] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[19])
                else:
                    listaConteo.append(lista_param_numerica[19])
                # listaConteo.append(5)
            if edad >= 18 and edad <= 25:
                if lista_param_numerica[20] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[20])
                else:
                    listaConteo.append(lista_param_numerica[20])
                # listaConteo.append(4)
            if edad >= 26 and edad <= 30:
                if lista_param_numerica[21] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[21])
                else:
                    listaConteo.append(lista_param_numerica[21])
                # listaConteo.append(5)
            if edad >= 31 and edad <= 35:
                if lista_param_numerica[22] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[22])
                else:
                    listaConteo.append(lista_param_numerica[22])
                # listaConteo.append(7)
            if edad >= 36 and edad <= 40:
                if lista_param_numerica[23] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[23])
                else:
                    listaConteo.append(lista_param_numerica[23])
                # listaConteo.append(9)
            if edad >= 41:
                if lista_param_numerica[24] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[24])
                else:
                    listaConteo.append(lista_param_numerica[24])
                # listaConteo.append(10)
            if trabajaConDependencia == "S":
                if lista_param_numerica[25] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[25])
                else:
                    listaConteo.append(lista_param_numerica[25])
                # listaConteo.append(10)
            if trabajaConDependencia == "N":
                if lista_param_numerica[26] == None or lista_param_numerica[0].isnumeric()==False:
                    listaConteo.append(listaParametros[26])
                else:
                    listaConteo.append(lista_param_numerica[26])
                # listaConteo.append(5)

        contador = 0
        listaConteoNum = [int(x) for x in listaConteo]
        for j in listaConteoNum:
            contador = contador + j
        puntaje.append(contador)  # hago que el puntaje se actualize al valor del puntaje que analizamos (del 0 al 100)
    dataFrameJT['PUNTAJE'] = pd.Series(puntaje)
    dataFrameFinal = dataFrameJT
    tupla = dataFrameFinal.shape
    numerof = tupla[0]
    listaderetorno = [dataFrameFinal, numerof]
    return (listaderetorno)  # devuelvo el dataframe con una columna extra donde va a estar los puntajes
# print(funcionPuntajes(dataFrameJT))

def existsFile(listaderetorno,allPath):
    dataFrameFinal=listaderetorno[0]
    dataFrameF = listaderetorno[0]
    numeroFinalFilas=listaderetorno[1]
    nuevopath=allPath + '\\' + 'Archivo_Perfilado.xlsx'
    if os.path.exists(nuevopath):
        dFinal = dataFrameFinal
        book = openpyxl.load_workbook(nuevopath)
        writer = pd.ExcelWriter(nuevopath, engine='openpyxl')
        writer.book = book

        pandas.io.formats.excel.header_style = None

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        ws=book.worksheets[0]
        vs=book.worksheets[0]
        end_sheet = ws.max_row

        # dFinal.to_excel(writer, ws.title, startrow = end_sheet, index=False )
        # ws.delete_rows(end_sheet + 1)
        serie_MU = list(dataFrameFinal['CEDULA - RUC'].to_dict().keys())
        dic_df= dataFrameFinal["CEDULA - RUC"].to_dict()
        #for i in serie_MU:
        comprobacion=False
        for j in range(end_sheet):
            #print(dataFrameFinal["CEDULA - RUC"].to_dict())
            #serie_MU = list(dataFrameFinal['CEDULA - RUC'].to_dict().keys())
            #for j in range(end_sheet):
            for i in serie_MU:
                if str(vs.cell((j+1),2).value) in dataFrameFinal["CEDULA - RUC"].to_dict().values():
                        print("La cédula "+str(vs.cell((j+1),2).value)+" tiene match.")
                    #serie_MU= list(dataFrameFinal['CEDULA - RUC'].to_dict().keys())
                    #for i in serie_MU:
                        ws.cell(i+2, 1, dataFrameFinal.iat[i, 0])
                        # ws.cell(i+2, 2, dataFrameFinal.iat[i, 1])
                        # ws.cell(i+2, 3, dataFrameFinal.iat[i, 2])
                        ws.cell(i+2, 4, dataFrameFinal.iat[i, 3])
                        ws.cell(i+2, 5, dataFrameFinal.iat[i, 4])
                        ws.cell(i+2, 6, dataFrameFinal.iat[i, 5])
                        ws.cell(i+2, 7, dataFrameFinal.iat[i, 6])
                        ws.cell(i+2, 8, dataFrameFinal.iat[i, 7])
                        ws.cell(i+2, 9, dataFrameFinal.iat[i, 8])
                        ws.cell(i+2, 10, dataFrameFinal.iat[i, 9])
                        ws.cell(i+2, 11, dataFrameFinal.iat[i, 10])
                        ws.cell(i+2, 12, dataFrameFinal.iat[i, 11])
                        ws.cell(i+2, 13, dataFrameFinal.iat[i, 12])
                        ws.cell(i+2, 14, dataFrameFinal.iat[i, 13])
                        ws.cell(i+2, 15, dataFrameFinal.iat[i, 14])
                        ws.cell(i+2, 16, dataFrameFinal.iat[i, 15])
                        ws.cell(i+2, 17, dataFrameFinal.iat[i, 16])
                        ws.cell(i+2, 18, dataFrameFinal.iat[i, 17])
                        #print(dataFrameF)
                        #dataFrameF = dFinal.drop(dFinal[dFinal["CEDULA - RUC"]!= vs.cell((j+1),2).value].index)
                        #print(dataFrameF)
                        #dataFrameF.to_excel(writer, ws.title, startrow=end_sheet, index=False)
                        #ws.delete_rows(end_sheet+1)
                #elif str(vs.cell((j+1),2).value) not in dataFrameFinal["CEDULA - RUC"].to_dict().values():
                        if j in dic_df.keys():
                            dataFrameF = dFinal.drop(dFinal[dFinal["CEDULA - RUC"] != dic_df[j]].index)
                            print(dataFrameF)
                            comprobacion = True
                            dimeF= dataFrameF.shape
                            for w in range(dimeF[0]):
                                if str(ws.cell((j + 1), 2).value) != dataFrameF.iat[w,1]:
                                    dataFrameF.to_excel(writer, ws.title, startrow=end_sheet, index=False)
                                    ws.delete_rows(end_sheet + 1)
                                    #writer.save()
                                    #nuevo_max=ws.max_row
                                    #print(nuevo_max)
                                    #ws.delete_rows(nuevo_max)
                elif str(vs.cell((j+1),2).value) not in dataFrameFinal["CEDULA - RUC"].to_dict().values() and comprobacion==False:
                    dFinal.to_excel(writer, ws.title, startrow=end_sheet, index=False)
                    ws.delete_rows(end_sheet + 1)
        #print(nuevo_max)
        writer.save()

        print("si") # no está entrando a esta opción

    else:
        print("no") # siempre a esta opcion

        createFile(dataFrameFinal,allPath)

def createFile(dataFrameFinal,allpath):
    print(dataFrameFinal[0])
    dFinal = dataFrameFinal[0]

    writer = ExcelWriter(allpath+"\\"+'Archivo_Perfilado.xlsx')
    dFinal.to_excel(writer, 'Hoja de datos', index=False)
    writer.save()
